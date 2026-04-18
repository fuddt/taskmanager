"""
Excel -> PlantUML Gantt Generator
PlantUML gantt 構文修正版

目的:
- SSOT Excel を読み込んで PlantUML の gantt 記法を生成する
- task / milestone / dependency を安定して出力する
- PlantUML 1.2024.x 系でも通りやすい形に寄せる

実行例:
python generate_plantuml_from_excel.py

出力:
- progress_gantt.puml
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


INPUT_FILE = "progress_template_plantuml_ssot.xlsx"
OUTPUT_FILE = "progress_gantt.puml"
SHEET_NAME = "進捗一覧"
HEADER_ROW = 6
DATA_START_ROW = 7


@dataclass
class TaskRow:
    task_id: str
    function_name: str
    task_category: str
    task_name: str
    display_name: str
    owner: str
    start_date: date
    end_date: date
    progress: int
    status: str
    priority: str
    display_order: int
    depends_on: List[str]
    task_type: str
    delay: str
    concern: str
    next_action: str
    updated_at: Optional[date]


def normalize_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_date(value) -> Optional[date]:
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    return None


def parse_int(value, default: int = 0) -> int:
    if value is None or value == "":
        return default

    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


def parse_depends(depends_text: str) -> List[str]:
    if not depends_text:
        return []

    items = [part.strip() for part in depends_text.split(",")]
    return [x for x in items if x]


def format_date_for_plantuml(d: date) -> str:
    return d.strftime("%Y-%m-%d")


def sanitize_label(text: str) -> str:
    """
    PlantUML gantt のタスク名として安全寄りの文字列に整形する

    方針:
    - 空白は _
    - 改行除去
    - [ ] は除去
    - 記号は極力減らす
    """
    s = normalize_str(text).replace("\n", " ").strip()

    # 角括弧は gantt 構文を壊しやすいので除去
    s = s.replace("[", "").replace("]", "")

    # 空白はアンダースコアへ
    s = s.replace(" ", "_")

    return s if s else "NO_NAME"


def build_unique_label(task: TaskRow) -> str:
    """
    表示名 + ID を安全なラベルにする

    例:
    A_ID_1
    """
    base = sanitize_label(task.display_name or task.task_name)
    return f"{base}_ID_{task.task_id}"


def build_header_map(ws) -> Dict[str, int]:
    header_map: Dict[str, int] = {}

    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=HEADER_ROW, column=col).value
        key = normalize_str(value)
        if key:
            header_map[key] = col

    return header_map


def required_columns() -> List[str]:
    return [
        "ID",
        "機能",
        "タスク区分",
        "タスク名",
        "表示名",
        "担当",
        "開始日",
        "期限",
        "進捗率",
        "ステータス",
        "優先度",
        "表示順",
        "依存タスクID",
        "タスク種別",
        "遅延",
        "課題/懸念",
        "次アクション",
        "更新日",
    ]


def validate_required_columns(header_map: Dict[str, int]) -> None:
    missing = [col for col in required_columns() if col not in header_map]
    if missing:
        raise ValueError("必須列が見つかりません: " + ", ".join(missing))


def is_empty_row(ws, row_idx: int, header_map: Dict[str, int]) -> bool:
    keys = ["ID", "タスク名", "開始日", "期限"]
    for key in keys:
        col = header_map[key]
        value = ws.cell(row=row_idx, column=col).value
        if value not in (None, ""):
            return False
    return True


def read_tasks_from_excel(file_path: str) -> List[TaskRow]:
    wb = load_workbook(file_path, data_only=False)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"シート '{SHEET_NAME}' が見つかりません")

    ws = wb[SHEET_NAME]
    header_map = build_header_map(ws)
    validate_required_columns(header_map)

    tasks: List[TaskRow] = []

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        if is_empty_row(ws, row_idx, header_map):
            continue

        task_id = normalize_str(ws.cell(row=row_idx, column=header_map["ID"]).value)
        function_name = normalize_str(ws.cell(row=row_idx, column=header_map["機能"]).value)
        task_category = normalize_str(ws.cell(row=row_idx, column=header_map["タスク区分"]).value)
        task_name = normalize_str(ws.cell(row=row_idx, column=header_map["タスク名"]).value)
        display_name = normalize_str(ws.cell(row=row_idx, column=header_map["表示名"]).value)
        owner = normalize_str(ws.cell(row=row_idx, column=header_map["担当"]).value)

        start_date = parse_date(ws.cell(row=row_idx, column=header_map["開始日"]).value)
        end_date = parse_date(ws.cell(row=row_idx, column=header_map["期限"]).value)

        progress = parse_int(ws.cell(row=row_idx, column=header_map["進捗率"]).value, default=0)
        status = normalize_str(ws.cell(row=row_idx, column=header_map["ステータス"]).value)
        priority = normalize_str(ws.cell(row=row_idx, column=header_map["優先度"]).value)
        display_order = parse_int(ws.cell(row=row_idx, column=header_map["表示順"]).value, default=9999)

        depends_text = normalize_str(ws.cell(row=row_idx, column=header_map["依存タスクID"]).value)
        depends_on = parse_depends(depends_text)

        task_type = normalize_str(ws.cell(row=row_idx, column=header_map["タスク種別"]).value).lower()
        delay = normalize_str(ws.cell(row=row_idx, column=header_map["遅延"]).value)
        concern = normalize_str(ws.cell(row=row_idx, column=header_map["課題/懸念"]).value)
        next_action = normalize_str(ws.cell(row=row_idx, column=header_map["次アクション"]).value)
        updated_at = parse_date(ws.cell(row=row_idx, column=header_map["更新日"]).value)

        if not task_id:
            raise ValueError(f"{row_idx}行目: ID が空です")

        if not task_name:
            raise ValueError(f"{row_idx}行目: タスク名 が空です (ID={task_id})")

        if not display_name:
            display_name = task_name

        if start_date is None:
            raise ValueError(f"{row_idx}行目: 開始日 が不正または空です (ID={task_id})")

        if end_date is None:
            raise ValueError(f"{row_idx}行目: 期限 が不正または空です (ID={task_id})")

        if not task_type:
            task_type = "task"

        if task_type not in ("task", "milestone"):
            raise ValueError(
                f"{row_idx}行目: タスク種別 は task / milestone のみ対応です "
                f"(ID={task_id}, 値={task_type})"
            )

        tasks.append(
            TaskRow(
                task_id=task_id,
                function_name=function_name,
                task_category=task_category,
                task_name=task_name,
                display_name=display_name,
                owner=owner,
                start_date=start_date,
                end_date=end_date,
                progress=progress,
                status=status,
                priority=priority,
                display_order=display_order,
                depends_on=depends_on,
                task_type=task_type,
                delay=delay,
                concern=concern,
                next_action=next_action,
                updated_at=updated_at,
            )
        )

    return tasks


def find_project_start(tasks: List[TaskRow]) -> date:
    return min(task.start_date for task in tasks)


def build_note_comment(task: TaskRow) -> List[str]:
    lines: List[str] = []

    meta = [
        f"ID={task.task_id}",
        f"機能={task.function_name}",
        f"区分={task.task_category}",
        f"担当={task.owner}",
        f"進捗率={task.progress}",
        f"状態={task.status}",
        f"優先度={task.priority}",
    ]

    if task.delay:
        meta.append(f"遅延={task.delay}")

    lines.append("' " + " / ".join(meta))

    if task.concern:
        lines.append(f"' 課題/懸念: {task.concern}")

    if task.next_action:
        lines.append(f"' 次アクション: {task.next_action}")

    return lines


def build_task_definition_lines(task: TaskRow, label_map: Dict[str, str]) -> List[str]:
    """
    タスク定義行を生成する

    ここでは alias を使わず、[label] そのものを参照する
    """
    lines: List[str] = []

    label = label_map[task.task_id]

    if task.task_type == "milestone":
        event_date = format_date_for_plantuml(task.end_date)
        lines.append(f"[{label}] happens {event_date}")
    else:
        start_str = format_date_for_plantuml(task.start_date)
        end_str = format_date_for_plantuml(task.end_date)
        lines.append(f"[{label}] starts {start_str}")
        lines.append(f"[{label}] ends {end_str}")

    return lines


def build_dependency_lines(task: TaskRow, label_map: Dict[str, str]) -> List[str]:
    """
    依存関係を生成する

    例:
    [詳細設計 [ID:2]] starts after [要件整理 [ID:1]]'s end
    """
    lines: List[str] = []
    current_label = label_map[task.task_id]

    for dep_id in task.depends_on:
        if dep_id not in label_map:
            lines.append(f"' WARNING: dependency not found: {task.task_id} depends on {dep_id}")
            continue

        dep_label = label_map[dep_id]
        lines.append(f"[{current_label}] starts after [{dep_label}]'s end")

    return lines


def generate_plantuml(tasks: List[TaskRow]) -> str:
    if not tasks:
        raise ValueError("有効なタスクが1件もありません")

    tasks = sorted(tasks, key=lambda x: (x.display_order, x.start_date, x.task_id))

    project_start = find_project_start(tasks)

    # task_id -> 一意ラベル
    label_map: Dict[str, str] = {}
    for task in tasks:
        label_map[task.task_id] = build_unique_label(task)

    lines: List[str] = []
    lines.append("@startgantt")
    lines.append("")
    lines.append("title Progress Gantt")
    lines.append(f"Project starts {format_date_for_plantuml(project_start)}")
    lines.append("saturday are closed")
    lines.append("sunday are closed")
    lines.append("")

    current_function = None

    for task in tasks:
        if task.function_name != current_function:
            current_function = task.function_name
            lines.append("' ========================================")
            lines.append(f"' 機能: {current_function or '未分類'}")
            lines.append("' ========================================")

        lines.extend(build_note_comment(task))
        lines.extend(build_task_definition_lines(task, label_map))
        lines.extend(build_dependency_lines(task, label_map))
        lines.append("")

    lines.append("@endgantt")
    lines.append("")

    return "\n".join(lines)


def write_text_file(file_path: str, content: str) -> None:
    Path(file_path).write_text(content, encoding="utf-8")


def main() -> None:
    tasks = read_tasks_from_excel(INPUT_FILE)
    puml_text = generate_plantuml(tasks)
    write_text_file(OUTPUT_FILE, puml_text)

    print(f"PlantUML file generated: {OUTPUT_FILE}")
    print(f"Task count: {len(tasks)}")


if __name__ == "__main__":
    main()