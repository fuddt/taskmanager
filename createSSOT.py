"""
SSOT Excel Generator v2
PlantUML生成対応版

目的:
- 進捗管理Excelを完全自動生成する
- 将来的に Excel -> PlantUML 変換しやすい構造にする
- GitHub 上でテンプレ仕様を管理できるようにする

実行:
python create_ssot_excel.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule


# =========================
# 共通設定
# =========================

# 進捗一覧シート名
SHEET_PROGRESS = "進捗一覧"

# マスタシート名
SHEET_MASTER = "マスタ"

# 週次サマリシート名
SHEET_WEEKLY = "週次サマリ"

# 課題・リスクシート名
SHEET_RISK = "課題・リスク"


def apply_header_style(cell):
    """
    ヘッダセル用の簡易スタイルを適用する
    """
    cell.font = Font(bold=True)
    cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def set_column_widths(ws, widths):
    """
    指定した列幅を設定する

    widths:
        {
            "A": 10,
            "B": 20,
            ...
        }
    """
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_dropdown(ws, cell_range, formula):
    """
    ドロップダウンを追加する

    例:
        add_dropdown(ws, "A2:A100", "=マスタ!$A$2:$A$5")
    """
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_integer_validation(ws, cell_range, min_value, max_value):
    """
    整数入力バリデーションを追加する
    """
    dv = DataValidation(
        type="whole",
        operator="between",
        formula1=str(min_value),
        formula2=str(max_value),
        allow_blank=True
    )
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_date_validation(ws, cell_range):
    """
    日付入力バリデーションを追加する
    """
    dv = DataValidation(type="date", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def freeze_header(ws, cell_ref):
    """
    ヘッダ固定
    """
    ws.freeze_panes = cell_ref


# =========================
# マスタシート作成
# =========================

def create_master(ws):
    """
    各種ドロップダウン候補を持つマスタシートを作成する

    このシートはユーザーが日常的に編集する想定ではなく、
    テンプレ側の入力規則を支える参照元として使う。
    """
    headers = [
        "ステータス", "優先度", "影響度", "タスク種別", "担当者候補", "タスク区分候補"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        apply_header_style(cell)

    status_list = ["未着手", "進行中", "完了", "保留"]
    priority_list = ["高", "中", "低"]
    impact_list = ["高", "中", "低"]

    # PlantUML変換に使う想定の種別
    # task: 通常タスク
    # milestone: 会議、承認、レビュー完了などの点イベント
    task_type_list = ["task", "milestone"]

    # 必要に応じて増やせるよう、ひとまず例を入れておく
    owner_list = ["自分", "部下A", "部下B"]
    task_category_list = ["要件定義", "設計", "実装", "テスト", "調整", "会議", "その他"]

    for i, value in enumerate(status_list, start=2):
        ws[f"A{i}"] = value

    for i, value in enumerate(priority_list, start=2):
        ws[f"B{i}"] = value

    for i, value in enumerate(impact_list, start=2):
        ws[f"C{i}"] = value

    for i, value in enumerate(task_type_list, start=2):
        ws[f"D{i}"] = value

    for i, value in enumerate(owner_list, start=2):
        ws[f"E{i}"] = value

    for i, value in enumerate(task_category_list, start=2):
        ws[f"F{i}"] = value

    set_column_widths(ws, {
        "A": 14,
        "B": 12,
        "C": 12,
        "D": 14,
        "E": 16,
        "F": 16,
    })


# =========================
# 進捗一覧シート作成
# =========================

def create_progress(ws):
    """
    SSOT本体となる進捗一覧シートを作成する

    PlantUML変換に必要な追加列:
    - 表示名
    - 表示順
    - 依存タスクID
    - タスク種別

    進捗率は 0〜100 の整数に固定する
    """
    headers = [
        "ID",              # 一意なID
        "機能",            # 画面/機能単位のまとまり
        "タスク区分",      # 要件定義 / 設計 / 実装 / テスト ...
        "タスク名",        # 正式な長い名前
        "表示名",          # PlantUML上で短く見せたい名前
        "担当",            # 担当者
        "開始日",          # 開始日
        "期限",            # 終了予定日
        "進捗率",          # 0〜100 の整数
        "ステータス",      # 未着手/進行中/完了/保留
        "優先度",          # 高/中/低
        "表示順",          # 図に出すときの順序
        "依存タスクID",    # 例: 1 または 2,3
        "タスク種別",      # task / milestone
        "遅延",            # 自動計算
        "課題/懸念",       # 会議で確認したいメモ
        "次アクション",    # 次にやること
        "更新日",          # 最終更新日
    ]

    # 6行目にヘッダを置く
    header_row = 6
    data_start_row = 7
    data_end_row = 200

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        apply_header_style(cell)

    # 説明文
    ws["A1"] = "入力ルール"
    ws["B1"] = "青字セルを入力。進捗一覧をSSOTとして更新する。PlantUML生成時は表示順・依存タスクID・タスク種別を使用する。"

    ws["A2"] = "進捗率ルール"
    ws["B2"] = "進捗率は 0〜100 の整数で入力する。例: 0, 40, 70, 100"

    ws["A3"] = "依存タスクIDルール"
    ws["B3"] = "依存がなければ空欄。複数依存はカンマ区切りで入力。例: 2,3"

    ws["A4"] = "タスク種別ルール"
    ws["B4"] = "通常作業は task。レビュー会や仕様確定など点イベントは milestone。"

    # ドロップダウン
    add_dropdown(ws, f"J{data_start_row}:J{data_end_row}", f"={SHEET_MASTER}!$A$2:$A$5")
    add_dropdown(ws, f"K{data_start_row}:K{data_end_row}", f"={SHEET_MASTER}!$B$2:$B$4")
    add_dropdown(ws, f"N{data_start_row}:N{data_end_row}", f"={SHEET_MASTER}!$D$2:$D$3")
    add_dropdown(ws, f"F{data_start_row}:F{data_end_row}", f"={SHEET_MASTER}!$E$2:$E$4")
    add_dropdown(ws, f"C{data_start_row}:C{data_end_row}", f"={SHEET_MASTER}!$F$2:$F$8")

    # 進捗率は 0〜100 の整数
    add_integer_validation(ws, f"I{data_start_row}:I{data_end_row}", 0, 100)

    # 表示順も整数
    add_integer_validation(ws, f"L{data_start_row}:L{data_end_row}", 1, 9999)

    # 日付列
    add_date_validation(ws, f"G{data_start_row}:G{data_end_row}")
    add_date_validation(ws, f"H{data_start_row}:H{data_end_row}")
    add_date_validation(ws, f"R{data_start_row}:R{data_end_row}")

    # 遅延列は自動計算
    # 期限 < 今日 かつ ステータス != 完了 なら「遅延」
    for row in range(data_start_row, data_end_row + 1):
        ws[f"O{row}"] = f'=IF(AND(H{row}<TODAY(),J{row}<>"完了",H{row}<>""),"遅延","")'

    # 条件付き書式
    # 完了ならステータス列を薄緑
    ws.conditional_formatting.add(
        f"J{data_start_row}:J{data_end_row}",
        FormulaRule(
            formula=[f'J{data_start_row}="完了"'],
            fill=PatternFill(fill_type="solid", fgColor="E2F0D9")
        )
    )

    # 遅延なら遅延列を薄赤
    ws.conditional_formatting.add(
        f"O{data_start_row}:O{data_end_row}",
        FormulaRule(
            formula=[f'O{data_start_row}="遅延"'],
            fill=PatternFill(fill_type="solid", fgColor="FCE4D6")
        )
    )

    # 列幅
    set_column_widths(ws, {
        "A": 8,
        "B": 18,
        "C": 14,
        "D": 28,
        "E": 18,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 10,
        "J": 12,
        "K": 10,
        "L": 10,
        "M": 14,
        "N": 12,
        "O": 10,
        "P": 28,
        "Q": 24,
        "R": 12,
    })

    freeze_header(ws, "A7")


# =========================
# 週次サマリシート作成
# =========================

def create_weekly(ws):
    """
    会議用の簡易サマリシートを作成する
    """
    ws["A1"] = "週次サマリ"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "完了件数"
    ws["B3"] = f'=COUNTIF({SHEET_PROGRESS}!J:J,"完了")'

    ws["A4"] = "進行中件数"
    ws["B4"] = f'=COUNTIF({SHEET_PROGRESS}!J:J,"進行中")'

    ws["A5"] = "未着手件数"
    ws["B5"] = f'=COUNTIF({SHEET_PROGRESS}!J:J,"未着手")'

    ws["A6"] = "遅延件数"
    ws["B6"] = f'=COUNTIF({SHEET_PROGRESS}!O:O,"遅延")'

    ws["A8"] = "確認ポイント"
    ws["A9"] = "1. 全体進捗率が前回比で上がっているか"
    ws["A10"] = "2. 遅延件数が0でない場合、期限再調整が必要か"
    ws["A11"] = "3. 課題・リスクの対応期限が近いものを確認"

    set_column_widths(ws, {
        "A": 42,
        "B": 14,
    })


# =========================
# 課題・リスクシート作成
# =========================

def create_risk(ws):
    """
    課題・リスク管理シートを作成する

    進捗一覧との接続を少し強くするため、
    影響タスクID列を持たせる
    """
    headers = [
        "ID",
        "分類",
        "内容",
        "影響タスクID",
        "影響度",
        "優先度",
        "担当",
        "発生日",
        "対応期限",
        "対応状況",
        "対応内容",
        "更新日",
    ]

    header_row = 4
    data_start_row = 5
    data_end_row = 200

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        apply_header_style(cell)

    ws["A1"] = "使い方"
    ws["B1"] = "進捗を止めそうなものだけ記載。影響タスクIDで進捗一覧とのつながりを持たせる。"

    # ドロップダウン
    add_dropdown(ws, f"E{data_start_row}:E{data_end_row}", f"={SHEET_MASTER}!$C$2:$C$4")
    add_dropdown(ws, f"F{data_start_row}:F{data_end_row}", f"={SHEET_MASTER}!$B$2:$B$4")
    add_dropdown(ws, f"J{data_start_row}:J{data_end_row}", f"={SHEET_MASTER}!$A$2:$A$5")
    add_dropdown(ws, f"G{data_start_row}:G{data_end_row}", f"={SHEET_MASTER}!$E$2:$E$4")

    add_date_validation(ws, f"H{data_start_row}:H{data_end_row}")
    add_date_validation(ws, f"I{data_start_row}:I{data_end_row}")
    add_date_validation(ws, f"L{data_start_row}:L{data_end_row}")

    set_column_widths(ws, {
        "A": 8,
        "B": 14,
        "C": 32,
        "D": 14,
        "E": 10,
        "F": 10,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 28,
        "L": 12,
    })

    freeze_header(ws, "A5")


# =========================
# メイン処理
# =========================

def main():
    """
    ワークブックを生成して保存する
    """
    wb = Workbook()

    # 進捗一覧
    progress = wb.active
    progress.title = SHEET_PROGRESS

    # 補助シート
    master = wb.create_sheet(SHEET_MASTER)
    weekly = wb.create_sheet(SHEET_WEEKLY)
    risk = wb.create_sheet(SHEET_RISK)

    # 各シート生成
    create_master(master)
    create_progress(progress)
    create_weekly(weekly)
    create_risk(risk)

    # 保存
    output_file = "progress_template_plantuml_ssot.xlsx"
    wb.save(output_file)

    print(f"SSOT Excel generated: {output_file}")


if __name__ == "__main__":
    main()